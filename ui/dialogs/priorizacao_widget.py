# -*- coding: utf-8 -*-
"""
Widget da aba "Priorização" da janela de configuração da elegibilidade.

Apresenta:
  - Checkboxes para ativar/desativar cada critério (A1..A8 e P1..P3)
  - Campo para carregar planilha de candidatos (CSV/XLSX)
  - Diálogo de mapeamento de colunas da planilha (CPF, CAF, sexo, sociobio)
"""

import csv
import os
from typing import Dict, List, Optional

from qgis.PyQt.QtCore import Qt
from qgis.PyQt.QtWidgets import (
    QCheckBox, QComboBox, QDialog, QDialogButtonBox, QFileDialog,
    QFormLayout, QGroupBox, QHBoxLayout, QLabel, QLineEdit, QMessageBox,
    QPushButton, QVBoxLayout, QWidget,
)


CRITERIOS_AREA = [
    ("crit_a1_ativo", "A1", "Municípios prioritários (>50% área)"),
    ("crit_a2_ativo", "A2", "Municípios desmate sob controle (>50% área)"),
    ("crit_a3_ativo", "A3", "Municípios Programa União (>50% área)"),
    ("crit_a4_ativo", "A4", "Áreas prioritárias biodiversidade (intersect)"),
    ("crit_a5_ativo", "A5", "Entorno de UC — 3km (exceto APA/RPPN)"),
    ("crit_a6_ativo", "A6", "Sobreposto ≥50% com APA ou RPPN"),
    ("crit_a7_ativo", "A7", "Entorno de TI — 3km"),
    ("crit_a8_ativo", "A8", "Bioma Amazônia (IBGE 2019)"),
]

CRITERIOS_PROVEDOR = [
    ("prio_caf", "P1", "Inscrito no CAF/DAP-PRONAF"),
    ("prio_sexo_feminino", "P2", "Proprietária do sexo feminino"),
    ("prio_sociobio", "P3", "Produtor de sociobiodiversidade"),
]

CAMPOS_CANDIDATOS = [
    ("cpf", "CPF", True),
    ("caf", "CAF/DAP-PRONAF (P1)", False),
    ("sexo", "Sexo (P2)", False),
    ("sociobio", "Sociobiodiversidade (P3)", False),
]


class PriorizacaoWidget(QWidget):
    """Widget da aba Priorização."""

    def __init__(self, config_atual: Optional[Dict] = None, parent=None):
        super().__init__(parent)
        self._config_atual = config_atual or {}
        self._planilha_atual: str = self._config_atual.get("planilha_candidatos", "")
        self._mapeamento_candidatos: Dict[str, str] = dict(
            self._config_atual.get("mapeamento_candidatos", {}) or {}
        )
        self._construir_ui()
        self._aplicar_estado_inicial()

    def _construir_ui(self):
        layout = QVBoxLayout(self)

        info = QLabel(
            "<b>Critérios de Priorização</b> (Item 6 do Edital). Os critérios "
            "selecionados são avaliados em uma etapa adicional após a elegibilidade. "
            "O resultado é uma pontuação que ordena os imóveis elegíveis em ranking."
        )
        info.setStyleSheet("font-size: 9px; color: #444; padding: 4px;")
        info.setWordWrap(True)
        layout.addWidget(info)

        # Critérios de área
        grp_area = QGroupBox("Critérios de Área (peso 1 cada)")
        l_area = QVBoxLayout(grp_area)
        self._cbs_area: Dict[str, QCheckBox] = {}
        for chave, sigla, rotulo in CRITERIOS_AREA:
            cb = QCheckBox(f"[{sigla}]  {rotulo}")
            cb.setChecked(bool(self._config_atual.get(chave, True)))
            self._cbs_area[chave] = cb
            l_area.addWidget(cb)
        layout.addWidget(grp_area)

        # Critérios do provedor
        grp_prov = QGroupBox("Critérios do Provedor (peso 3 cada)")
        l_prov = QVBoxLayout(grp_prov)

        info_prov = QLabel(
            "Avaliados <b>somente se</b> uma planilha de candidatos for carregada. "
            "Caso a planilha não esteja disponível, o processamento continua "
            "normalmente sem esses critérios."
        )
        info_prov.setStyleSheet("font-size: 9px; color: #555;")
        info_prov.setWordWrap(True)
        l_prov.addWidget(info_prov)

        for chave, sigla, rotulo in CRITERIOS_PROVEDOR:
            lbl = QLabel(f"  • [{sigla}]  {rotulo}")
            lbl.setStyleSheet("color: #2c3e50;")
            l_prov.addWidget(lbl)

        # Linha da planilha
        planilha_layout = QHBoxLayout()
        planilha_layout.addWidget(QLabel("Planilha:"))
        self.edit_planilha = QLineEdit()
        self.edit_planilha.setReadOnly(True)
        self.edit_planilha.setPlaceholderText("(opcional) Carregar CSV/XLSX de candidatos...")
        planilha_layout.addWidget(self.edit_planilha, 1)

        self.btn_carregar = QPushButton("Carregar...")
        self.btn_carregar.clicked.connect(self._selecionar_planilha)
        planilha_layout.addWidget(self.btn_carregar)

        self.btn_limpar_planilha = QPushButton("Limpar")
        self.btn_limpar_planilha.clicked.connect(self._limpar_planilha)
        planilha_layout.addWidget(self.btn_limpar_planilha)

        l_prov.addLayout(planilha_layout)

        # Status do mapeamento
        self.lbl_mapeamento = QLabel()
        self.lbl_mapeamento.setStyleSheet("font-size: 9px; color: #555;")
        self.lbl_mapeamento.setWordWrap(True)
        l_prov.addWidget(self.lbl_mapeamento)

        layout.addWidget(grp_prov)
        layout.addStretch()

    def _aplicar_estado_inicial(self):
        if self._planilha_atual:
            self.edit_planilha.setText(self._planilha_atual)
        self._atualizar_status_mapeamento()

    def _atualizar_status_mapeamento(self):
        if not self._planilha_atual:
            self.lbl_mapeamento.setText("<i>Nenhuma planilha carregada — P1, P2 e P3 serão pulados.</i>")
            return
        if not self._mapeamento_candidatos:
            self.lbl_mapeamento.setText(
                "<span style='color: #c0392b;'>⚠ Mapeamento de colunas não configurado.</span>"
            )
            return
        partes = []
        for chave, rotulo, _ in CAMPOS_CANDIDATOS:
            valor = self._mapeamento_candidatos.get(chave)
            if valor:
                partes.append(f"{rotulo}: <b>{valor}</b>")
        self.lbl_mapeamento.setText("Mapeamento: " + " | ".join(partes))

    # ------------------------------------------------------------------ #
    # Planilha                                                              #
    # ------------------------------------------------------------------ #

    def _selecionar_planilha(self):
        caminho, _ = QFileDialog.getOpenFileName(
            self,
            "Selecionar planilha de candidatos",
            "",
            "Planilhas (*.csv *.xlsx *.xls);;CSV (*.csv);;Excel (*.xlsx *.xls)",
        )
        if not caminho:
            return

        # Detectar colunas da planilha
        colunas = self._detectar_colunas(caminho)
        if not colunas:
            QMessageBox.warning(
                self, "Planilha inválida",
                f"Não foi possível ler colunas da planilha:\n{caminho}",
            )
            return

        # Diálogo de mapeamento
        dlg = MapeamentoCandidatosDialog(colunas, self._mapeamento_candidatos, self)
        if dlg.exec_() != QDialog.Accepted:
            return
        novo_mapa = dlg.get_mapeamento()
        if not novo_mapa.get("cpf"):
            QMessageBox.warning(
                self, "CPF obrigatório",
                "É necessário mapear a coluna de CPF para usar a planilha.",
            )
            return

        self._planilha_atual = caminho
        self._mapeamento_candidatos = novo_mapa
        self.edit_planilha.setText(caminho)
        self._atualizar_status_mapeamento()

    def _limpar_planilha(self):
        self._planilha_atual = ""
        self._mapeamento_candidatos = {}
        self.edit_planilha.clear()
        self._atualizar_status_mapeamento()

    @staticmethod
    def _detectar_colunas(caminho: str) -> List[str]:
        ext = os.path.splitext(caminho)[1].lower()
        try:
            if ext == ".csv":
                with open(caminho, "r", encoding="utf-8-sig", newline="") as f:
                    sample = f.read(4096)
                    f.seek(0)
                    try:
                        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                    except Exception:
                        class _D(csv.excel):
                            delimiter = ";"
                        dialect = _D
                    reader = csv.reader(f, dialect=dialect)
                    header = next(reader, [])
                    return [h.strip() for h in header if h]
            if ext in (".xlsx", ".xls"):
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    return []
                wb = load_workbook(caminho, read_only=True, data_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                header = next(rows_iter, ())
                return [str(c).strip() for c in header if c is not None]
        except Exception:
            return []
        return []

    # ------------------------------------------------------------------ #
    # API pública                                                          #
    # ------------------------------------------------------------------ #

    def get_config_priorizacao(self) -> Dict:
        """Retorna dict com config a ser aplicada no ConfigProcessamento."""
        cfg = {chave: cb.isChecked() for chave, cb in self._cbs_area.items()}
        cfg["planilha_candidatos"] = self._planilha_atual
        cfg["mapeamento_candidatos"] = dict(self._mapeamento_candidatos)
        return cfg


class MapeamentoCandidatosDialog(QDialog):
    """Diálogo para mapear colunas da planilha de candidatos."""

    def __init__(self, colunas: List[str], mapeamento_atual: Dict[str, str], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Mapeamento de colunas — Planilha de Candidatos")
        self.setMinimumWidth(500)

        self._colunas = ["(nenhuma)"] + sorted(colunas)
        self._mapeamento_atual = dict(mapeamento_atual or {})
        self._combos: Dict[str, QComboBox] = {}

        layout = QVBoxLayout(self)
        info = QLabel(
            "Selecione qual coluna da planilha corresponde a cada campo. "
            "<b>CPF</b> é obrigatório. Os demais são opcionais — critérios "
            "sem coluna mapeada serão considerados \"Não\"."
        )
        info.setStyleSheet("font-size: 10px; color: #444;")
        info.setWordWrap(True)
        layout.addWidget(info)

        form = QFormLayout()
        # Auto-detectar mapeamentos óbvios pelo nome da coluna
        sugestoes = self._sugerir_mapeamentos()

        for chave, rotulo, obrigatorio in CAMPOS_CANDIDATOS:
            cb = QComboBox()
            cb.addItems(self._colunas)
            valor_atual = self._mapeamento_atual.get(chave) or sugestoes.get(chave)
            if valor_atual and valor_atual in self._colunas:
                cb.setCurrentText(valor_atual)
            else:
                cb.setCurrentIndex(0)
            self._combos[chave] = cb
            label = f"{rotulo}{' *' if obrigatorio else ''}:"
            form.addRow(label, cb)
        layout.addLayout(form)

        nota = QLabel("<small>* Campo obrigatório</small>")
        layout.addWidget(nota)

        botoes = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        botoes.accepted.connect(self.accept)
        botoes.rejected.connect(self.reject)
        layout.addWidget(botoes)

    def _sugerir_mapeamentos(self) -> Dict[str, str]:
        """Tenta auto-detectar colunas pelos nomes."""
        sugestoes = {}
        candidates = {
            "cpf": ["cpf", "cpf_cnpj", "documento", "doc", "cnpj_cpf"],
            "caf": ["caf", "dap", "dap_pronaf", "caf_dap", "pronaf"],
            "sexo": ["sexo", "genero", "sexo_proprietario"],
            "sociobio": [
                "sociobio", "sociobiodiversidade", "extrativismo",
                "extrativista", "babacu", "andiroba",
            ],
        }
        cols_lower = {c.lower(): c for c in self._colunas}
        for chave, opts in candidates.items():
            for opt in opts:
                if opt in cols_lower:
                    sugestoes[chave] = cols_lower[opt]
                    break
        return sugestoes

    def get_mapeamento(self) -> Dict[str, str]:
        resultado = {}
        for chave, cb in self._combos.items():
            valor = cb.currentText()
            if valor and valor != "(nenhuma)":
                resultado[chave] = valor
        return resultado
