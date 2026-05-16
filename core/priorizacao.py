# -*- coding: utf-8 -*-
"""
Módulo de Análise de Priorização para o Plugin Floresta+ Amazônia.

Implementa os critérios de priorização do Item 6 do Edital
(Chamada Pública 02/2024 — Modalidade Conservação):

  Critérios de Área (peso 1 cada):
    A1 — >50% em municípios prioritários
    A2 — >50% em municípios c/ desmatamento monitorado e sob controle
    A3 — >50% em municípios do Programa União com Municípios
    A4 — >50% em municípios c/ áreas prioritárias para biodiversidade
    A5 — Entorno de UC (3km), exceto APA e RPPN
    A6 — Sobreposto ≥50% com APA ou RPPN
    A7 — Entorno de TI/PIPCT (3km)
    A8 — Localizado no bioma Amazônia (IBGE)

  Critérios do Provedor (peso 3 cada):
    P1 — Inscrito no CAF/DAP-PRONAF
    P2 — Proprietária do sexo feminino
    P3 — Produtor de sociobiodiversidade

Importante: Este módulo é uma análise ADICIONAL — não interfere com a
elegibilidade existente. É executado como uma etapa extra ao final do
processamento, após a elegibilidade já estar calculada.

Autor: Plugin Floresta+
Data: Maio 2026
"""

import csv
import os
import traceback
from typing import Dict, List, Optional, Set, Tuple

try:
    from qgis.core import (
        QgsCoordinateReferenceSystem, QgsCoordinateTransform,
        QgsDistanceArea, QgsFeature, QgsField, QgsGeometry,
        QgsProject, QgsSpatialIndex, QgsUnitTypes, QgsVectorLayer,
    )
    from qgis.PyQt.QtCore import QVariant
except ImportError:
    pass


# Campos novos adicionados pela priorização
CAMPOS_PRIORIZACAO = [
    ("prio_mun_prioritario", "String", "Sim/Não — >50% área em municípios prioritários (A1)"),
    ("prio_mun_controle", "String", "Sim/Não — >50% área em mun. desmate sob controle (A2)"),
    ("prio_mun_uniao", "String", "Sim/Não — >50% área em mun. do Programa União (A3)"),
    ("prio_biodiversidade", "String", "Sim/Não — >50% área em mun. c/ áreas prioritárias (A4)"),
    ("prio_entorno_uc", "String", "Sim/Não — entorno 3km de UC (exceto APA/RPPN) (A5)"),
    ("prio_apa_rppn", "String", "Sim/Não — sobreposto ≥50% com APA ou RPPN (A6)"),
    ("prio_entorno_ti", "String", "Sim/Não — entorno 3km de TI (A7)"),
    ("prio_bioma_amazonia", "String", "Sim/Não — no bioma Amazônia IBGE (A8)"),
    ("prio_caf", "String", "Sim/Não — inscrito no CAF/DAP-PRONAF (P1)"),
    ("prio_sexo_feminino", "String", "Sim/Não — proprietária do sexo feminino (P2)"),
    ("prio_sociobio", "String", "Sim/Não — produtor de sociobiodiversidade (P3)"),
    ("score_priorizacao", "Integer", "Soma dos pesos dos critérios atendidos"),
    ("ranking", "Integer", "Posição no ranking (1 = mais prioritário)"),
    ("pct_rvn_total", "Double", "% RVN / área do imóvel (critério de desempate)"),
]


def _get_field_type(tipo_str: str):
    """Retorna o QVariant correspondente ao tipo string."""
    if tipo_str == "String":
        return QVariant.String
    if tipo_str == "Integer":
        return QVariant.Int
    if tipo_str == "Double":
        return QVariant.Double
    return QVariant.String


class PriorizacaoAnalise:
    """
    Análise de priorização — executada como etapa extra após a elegibilidade.

    Recebe a camada de saída da ProcessamentoElegiveis (já com colunas de
    elegibilidade preenchidas) e adiciona colunas de priorização.
    """

    def __init__(self, config, layer, callback_log=None, helper_carregar=None):
        """
        Args:
            config: ConfigProcessamento — configurações do processamento.
            layer: QgsVectorLayer — camada de elegíveis (em memória).
            callback_log: função (msg) para log.
            helper_carregar: função (nome_camada) para carregar camadas do GPKG.
                             Se None, tentará carregar diretamente do gpkg_path.
        """
        self.config = config
        self.layer = layer
        self.callback_log = callback_log
        self._helper_carregar = helper_carregar

        # Resultados intermediários por fid
        self._resultado: Dict[int, Dict[str, str]] = {}
        self._area_imovel: Dict[int, float] = {}  # área em ha do imóvel

        # CRS para cálculos de área
        try:
            self.crs_area = QgsCoordinateReferenceSystem(config.crs_area)
        except Exception:
            self.crs_area = QgsCoordinateReferenceSystem("ESRI:102033")

    # ------------------------------------------------------------------ #
    # Utilitários                                                          #
    # ------------------------------------------------------------------ #

    def log(self, msg: str):
        print(f"[Priorização] {msg}")
        if self.callback_log:
            self.callback_log(msg)

    def _carregar_camada(self, nome_camada: str) -> Optional["QgsVectorLayer"]:
        """Carrega camada do GeoPackage (delegando se possível)."""
        if self._helper_carregar:
            return self._helper_carregar(nome_camada)
        gpkg = getattr(self.config, "gpkg_path", "")
        if not gpkg or not os.path.exists(gpkg):
            return None
        uri = f"{gpkg}|layername={nome_camada}"
        layer = QgsVectorLayer(uri, nome_camada, "ogr")
        if not layer.isValid():
            return None
        return layer

    def _calcular_area_ha(self, geom: "QgsGeometry", crs_origem) -> float:
        """Calcula área de uma geometria em hectares."""
        if geom is None or geom.isEmpty():
            return 0.0
        da = QgsDistanceArea()
        da.setSourceCrs(crs_origem, QgsProject.instance().transformContext())
        da.setEllipsoid("WGS84")
        area_m2 = da.measureArea(geom)
        return da.convertAreaMeasurement(area_m2, QgsUnitTypes.AreaHectares)

    def _campo_existe(self, nome: str) -> bool:
        """Verifica se um campo existe na camada."""
        return self.layer.fields().indexOf(nome) >= 0

    def _adicionar_campos(self):
        """Adiciona os campos de priorização à camada (se não existirem)."""
        provider = self.layer.dataProvider()
        novos = []
        for nome, tipo, _ in CAMPOS_PRIORIZACAO:
            if not self._campo_existe(nome):
                novos.append(QgsField(nome, _get_field_type(tipo)))
        if novos:
            provider.addAttributes(novos)
            self.layer.updateFields()

    def _identificar_campo(self, nomes_possiveis: List[str], layer=None) -> Optional[str]:
        """Encontra o nome real (case-insensitive) de um campo na camada."""
        layer = layer or self.layer
        campos = [f.name() for f in layer.fields()]
        campos_lower = {c.lower(): c for c in campos}
        for nome in nomes_possiveis:
            if nome.lower() in campos_lower:
                return campos_lower[nome.lower()]
        return None

    # ------------------------------------------------------------------ #
    # Critérios de área baseados em listas de municípios (A1, A2, A3, A4) #
    # ------------------------------------------------------------------ #

    def _calcular_municipios_intersect(
        self,
        geocodigos: Set[str],
        layer_mun: "QgsVectorLayer",
        index_mun: "QgsSpatialIndex",
        campo_geo: str,
    ) -> Dict[int, bool]:
        """
        Para cada imóvel, calcula se >50% da área está dentro dos municípios
        cujos geocódigos estão no set 'geocodigos'.

        Retorna {fid: True/False}.
        """
        resultado: Dict[int, bool] = {}
        if not geocodigos:
            return resultado

        threshold = float(getattr(self.config, "threshold_area_pct", 0.50))

        for feat in self.layer.getFeatures():
            fid = feat.id()
            geom = feat.geometry()
            if geom is None or geom.isEmpty():
                resultado[fid] = False
                continue

            area_total = self._area_imovel.get(fid)
            if area_total is None:
                area_total = self._calcular_area_ha(geom, self.layer.crs())
                self._area_imovel[fid] = area_total

            if area_total <= 0:
                resultado[fid] = False
                continue

            # Iterar pelos municípios candidatos
            area_intersect = 0.0
            candidatos = index_mun.intersects(geom.boundingBox())
            for fid_mun in candidatos:
                feat_mun = layer_mun.getFeature(fid_mun)
                geo = str(feat_mun.attribute(campo_geo) or "").strip()
                if geo not in geocodigos:
                    continue
                geom_mun = feat_mun.geometry()
                if not geom.intersects(geom_mun):
                    continue
                inter = geom.intersection(geom_mun)
                if inter.isEmpty():
                    continue
                area_intersect += self._calcular_area_ha(inter, self.layer.crs())

            pct = area_intersect / area_total if area_total > 0 else 0.0
            resultado[fid] = pct > threshold

        return resultado

    def _executar_criterios_municipios(self):
        """Executa A1, A2, A3 e prepara dados para A4."""
        ativos = []
        if getattr(self.config, "crit_a1_ativo", True):
            ativos.append(("a1", "geocodigos_prioritarios", "prio_mun_prioritario"))
        if getattr(self.config, "crit_a2_ativo", True):
            ativos.append(("a2", "geocodigos_desmate_controle", "prio_mun_controle"))
        if getattr(self.config, "crit_a3_ativo", True):
            ativos.append(("a3", "geocodigos_programa_uniao", "prio_mun_uniao"))

        if not ativos and not getattr(self.config, "crit_a4_ativo", True):
            return

        layer_mun = self._carregar_camada(self.config.camada_municipios)
        if not layer_mun:
            self.log("⚠ Camada de municípios não encontrada — pulando A1/A2/A3/A4")
            return

        campo_geo = self._identificar_campo(
            ["geocodigo", "cod_municipio", "codmun", "cd_mun", "cd_geocmu", "ibge"],
            layer=layer_mun,
        )
        if not campo_geo:
            self.log("⚠ Campo de geocódigo não encontrado em 'municipios' — pulando A1/A2/A3/A4")
            return

        index_mun = QgsSpatialIndex(layer_mun.getFeatures())

        for cid, atributo_cfg, coluna_saida in ativos:
            geocodigos = getattr(self.config, atributo_cfg, set()) or set()
            self.log(f"   {cid.upper()}: {len(geocodigos)} geocódigos configurados")
            resultado = self._calcular_municipios_intersect(
                geocodigos, layer_mun, index_mun, campo_geo
            )
            atendem = sum(1 for v in resultado.values() if v)
            self.log(f"   {cid.upper()}: {atendem} imóveis com >50% área em municípios listados")
            for fid, ok in resultado.items():
                self._resultado.setdefault(fid, {})[coluna_saida] = "Sim" if ok else "Não"

        # A4 — Áreas Prioritárias para Conservação (intersect espacial)
        if getattr(self.config, "crit_a4_ativo", True):
            self._executar_a4(layer_mun, index_mun, campo_geo)

    def _executar_a4(self, layer_mun, index_mun, campo_geo: str):
        """A4 — gera lista dinâmica de geocódigos via intersect Áreas Prio. x Municípios."""
        layer_areas = self._carregar_camada(self.config.camada_areas_prioritarias)
        if not layer_areas:
            self.log(
                "⚠ A4: camada 'areas_prioritarias_conservacao' não encontrada — "
                "critério desabilitado"
            )
            for feat in self.layer.getFeatures():
                self._resultado.setdefault(feat.id(), {})["prio_biodiversidade"] = "Não"
            return

        self.log(f"   A4: identificando municípios via intersect com áreas prioritárias...")

        index_areas = QgsSpatialIndex(layer_areas.getFeatures())
        geocodigos_a4: Set[str] = set()

        for feat_mun in layer_mun.getFeatures():
            geom_mun = feat_mun.geometry()
            if geom_mun is None or geom_mun.isEmpty():
                continue
            candidatos = index_areas.intersects(geom_mun.boundingBox())
            for fid_a in candidatos:
                feat_a = layer_areas.getFeature(fid_a)
                if geom_mun.intersects(feat_a.geometry()):
                    geo = str(feat_mun.attribute(campo_geo) or "").strip()
                    if geo:
                        geocodigos_a4.add(geo)
                    break

        self.log(f"   A4: {len(geocodigos_a4)} municípios com áreas prioritárias identificados")

        resultado = self._calcular_municipios_intersect(
            geocodigos_a4, layer_mun, index_mun, campo_geo
        )
        atendem = sum(1 for v in resultado.values() if v)
        self.log(f"   A4: {atendem} imóveis com >50% área nesses municípios")
        for fid, ok in resultado.items():
            self._resultado.setdefault(fid, {})["prio_biodiversidade"] = "Sim" if ok else "Não"

    # ------------------------------------------------------------------ #
    # Critérios espaciais sobre UCs e TIs (A5, A6, A7)                   #
    # ------------------------------------------------------------------ #

    def _executar_uc_ti(self):
        """Executa A5, A6, A7 — UCs e TIs."""
        if (
            not getattr(self.config, "crit_a5_ativo", True)
            and not getattr(self.config, "crit_a6_ativo", True)
            and not getattr(self.config, "crit_a7_ativo", True)
        ):
            return

        # UCs (A5 e A6)
        if getattr(self.config, "crit_a5_ativo", True) or getattr(
            self.config, "crit_a6_ativo", True
        ):
            self._executar_uc()

        # TIs (A7)
        if getattr(self.config, "crit_a7_ativo", True):
            self._executar_ti()

    def _executar_uc(self):
        """A5 — entorno 3km UC (exceto APA/RPPN); A6 — ≥50% sobreposto APA/RPPN."""
        layer_uc = self._carregar_camada(self.config.camada_ucs)
        if not layer_uc:
            self.log("⚠ A5/A6: camada 'ucs' não encontrada — critérios desabilitados")
            for feat in self.layer.getFeatures():
                self._resultado.setdefault(feat.id(), {}).update({
                    "prio_entorno_uc": "Não",
                    "prio_apa_rppn": "Não",
                })
            return

        # Identificar campo de categoria
        campo_cat = self._identificar_campo(
            ["categoria", "cat_uc", "categoria_uc", "categori3", "categoria_legal"],
            layer=layer_uc,
        )
        if not campo_cat:
            self.log("⚠ A5/A6: campo categoria não encontrado em 'ucs' — usando todas as feições")

        # Buffer 3km — calcular em CRS projetado para garantir distância em metros
        buffer_uc_m = float(getattr(self.config, "buffer_uc_km", 3.0)) * 1000.0
        threshold_apa = float(getattr(self.config, "threshold_apa_rppn", 0.50))

        # Index com features filtradas
        # Para A5: UCs EXCETO APA/RPPN
        # Para A6: APENAS APA/RPPN
        feats_uc_geral = []  # (fid, geom_em_crs_imovel, categoria)
        for feat in layer_uc.getFeatures():
            geom = feat.geometry()
            if geom is None or geom.isEmpty():
                continue
            cat = ""
            if campo_cat:
                cat = str(feat.attribute(campo_cat) or "").strip()
            feats_uc_geral.append((feat.id(), geom, cat))

        # A5 — buffer 3km de UCs não-APA/RPPN
        if getattr(self.config, "crit_a5_ativo", True):
            self._aplicar_entorno(
                feats_uc_geral,
                layer_uc.crs(),
                buffer_uc_m,
                "prio_entorno_uc",
                excluir_categorias=("APA", "RPPN"),
                rotulo="A5 — Entorno UC 3km",
            )

        # A6 — sobreposição ≥50% com APA/RPPN
        if getattr(self.config, "crit_a6_ativo", True):
            self._aplicar_sobreposicao(
                feats_uc_geral,
                layer_uc.crs(),
                threshold_apa,
                "prio_apa_rppn",
                incluir_categorias=("APA", "RPPN"),
                rotulo="A6 — Sobreposição APA/RPPN",
            )

    def _executar_ti(self):
        """A7 — entorno 3km TI."""
        layer_ti = self._carregar_camada(self.config.camada_terras_indigenas)
        if not layer_ti:
            self.log("⚠ A7: camada 'terras_indigenas' não encontrada — critério desabilitado")
            for feat in self.layer.getFeatures():
                self._resultado.setdefault(feat.id(), {})["prio_entorno_ti"] = "Não"
            return

        buffer_ti_m = float(getattr(self.config, "buffer_ti_km", 3.0)) * 1000.0

        feats_ti = []
        for feat in layer_ti.getFeatures():
            geom = feat.geometry()
            if geom is None or geom.isEmpty():
                continue
            feats_ti.append((feat.id(), geom, ""))

        self._aplicar_entorno(
            feats_ti,
            layer_ti.crs(),
            buffer_ti_m,
            "prio_entorno_ti",
            excluir_categorias=(),
            rotulo="A7 — Entorno TI 3km",
        )

    def _aplicar_entorno(self, feats, crs_origem, buffer_metros, coluna_saida,
                          excluir_categorias=(), rotulo: str = ""):
        """
        Aplica buffer (em CRS projetado) e verifica interseção com cada imóvel.

        feats: lista de (fid, geom, categoria) na CRS de origem da camada.
        """
        if rotulo:
            self.log(f"   {rotulo}: gerando buffer de {buffer_metros/1000:.1f} km...")

        # Filtrar por categoria (string contém)
        feats_filtrados = []
        for fid, geom, cat in feats:
            if excluir_categorias:
                cat_upper = cat.upper()
                if any(ex in cat_upper for ex in excluir_categorias):
                    continue
            feats_filtrados.append((fid, geom))

        if not feats_filtrados:
            self.log(f"   {rotulo}: nenhuma feição após filtro — todos NÃO")
            for feat in self.layer.getFeatures():
                self._resultado.setdefault(feat.id(), {})[coluna_saida] = "Não"
            return

        # Transformar para CRS projetado para buffer em metros
        crs_proj = self.crs_area
        transform = QgsCoordinateTransform(
            crs_origem, crs_proj, QgsProject.instance()
        )
        transform_inv = QgsCoordinateTransform(
            crs_proj, self.layer.crs(), QgsProject.instance()
        )

        # Construir geometrias em buffer e índice espacial
        geoms_buffer = []
        for fid, geom in feats_filtrados:
            try:
                g = QgsGeometry(geom)
                g.transform(transform)
                buf = g.buffer(buffer_metros, 8)
                buf.transform(transform_inv)
                if not buf.isEmpty():
                    geoms_buffer.append(buf)
            except Exception:
                continue

        # Verificar interseção com cada imóvel
        contador_atende = 0
        for feat in self.layer.getFeatures():
            fid_imovel = feat.id()
            geom_imovel = feat.geometry()
            if geom_imovel is None or geom_imovel.isEmpty():
                self._resultado.setdefault(fid_imovel, {})[coluna_saida] = "Não"
                continue

            atende = False
            bbox = geom_imovel.boundingBox()
            for buf in geoms_buffer:
                if not buf.boundingBox().intersects(bbox):
                    continue
                if geom_imovel.intersects(buf):
                    atende = True
                    break
            self._resultado.setdefault(fid_imovel, {})[coluna_saida] = "Sim" if atende else "Não"
            if atende:
                contador_atende += 1

        if rotulo:
            self.log(f"   {rotulo}: {contador_atende} imóveis atendem")

    def _aplicar_sobreposicao(self, feats, crs_origem, threshold, coluna_saida,
                                incluir_categorias=(), rotulo: str = ""):
        """Calcula % de sobreposição entre imóvel e feições, retorna Sim/Não."""
        if rotulo:
            self.log(f"   {rotulo}: calculando sobreposição (≥{int(threshold*100)}%)...")

        # Filtrar por categoria (deve conter alguma das incluídas)
        feats_filtrados = []
        for fid, geom, cat in feats:
            if incluir_categorias:
                cat_upper = cat.upper()
                if not any(inc in cat_upper for inc in incluir_categorias):
                    continue
            feats_filtrados.append((fid, geom))

        if not feats_filtrados:
            self.log(f"   {rotulo}: nenhuma feição após filtro — todos NÃO")
            for feat in self.layer.getFeatures():
                self._resultado.setdefault(feat.id(), {})[coluna_saida] = "Não"
            return

        # Reprojetar geometrias para CRS da camada de imóveis (caso difiram)
        if crs_origem != self.layer.crs():
            transform = QgsCoordinateTransform(
                crs_origem, self.layer.crs(), QgsProject.instance()
            )
            geoms_validas = []
            for fid, g in feats_filtrados:
                try:
                    gg = QgsGeometry(g)
                    gg.transform(transform)
                    if not gg.isEmpty():
                        geoms_validas.append(gg)
                except Exception:
                    continue
        else:
            geoms_validas = [g for _, g in feats_filtrados]

        contador = 0
        for feat in self.layer.getFeatures():
            fid_imovel = feat.id()
            geom_imovel = feat.geometry()
            if geom_imovel is None or geom_imovel.isEmpty():
                self._resultado.setdefault(fid_imovel, {})[coluna_saida] = "Não"
                continue

            area_total = self._area_imovel.get(fid_imovel)
            if area_total is None:
                area_total = self._calcular_area_ha(geom_imovel, self.layer.crs())
                self._area_imovel[fid_imovel] = area_total

            if area_total <= 0:
                self._resultado.setdefault(fid_imovel, {})[coluna_saida] = "Não"
                continue

            bbox = geom_imovel.boundingBox()
            area_inter = 0.0
            for g in geoms_validas:
                if not g.boundingBox().intersects(bbox):
                    continue
                if not geom_imovel.intersects(g):
                    continue
                inter = geom_imovel.intersection(g)
                if not inter.isEmpty():
                    area_inter += self._calcular_area_ha(inter, self.layer.crs())

            pct = area_inter / area_total if area_total > 0 else 0.0
            atende = pct >= threshold
            self._resultado.setdefault(fid_imovel, {})[coluna_saida] = "Sim" if atende else "Não"
            if atende:
                contador += 1

        if rotulo:
            self.log(f"   {rotulo}: {contador} imóveis atendem")

    # ------------------------------------------------------------------ #
    # A8 — Bioma Amazônia                                                 #
    # ------------------------------------------------------------------ #

    def _executar_bioma_amazonia(self):
        """A8 — verifica se imóvel intersecta feição 'Amazônia' na camada de biomas."""
        if not getattr(self.config, "crit_a8_ativo", True):
            return

        layer_biomas = self._carregar_camada(self.config.camada_biomas)
        if not layer_biomas:
            self.log("⚠ A8: camada 'biomas' não encontrada — critério desabilitado")
            for feat in self.layer.getFeatures():
                self._resultado.setdefault(feat.id(), {})["prio_bioma_amazonia"] = "Não"
            return

        campo_nome = self._identificar_campo(
            ["bioma", "nome", "nome_bioma", "nm_bioma"],
            layer=layer_biomas,
        )

        # Filtrar feições com nome contendo "Amazônia"
        geoms_amaz = []
        for feat in layer_biomas.getFeatures():
            nome = ""
            if campo_nome:
                nome = str(feat.attribute(campo_nome) or "").strip().upper()
            geom = feat.geometry()
            if geom is None or geom.isEmpty():
                continue
            # Se houver campo de nome, filtrar; senão, usar todas (assume só Amazônia)
            if campo_nome:
                if "AMAZ" not in nome:
                    continue
            geoms_amaz.append(geom)

        if not geoms_amaz:
            self.log("⚠ A8: nenhuma feição 'Amazônia' encontrada — todos NÃO")
            for feat in self.layer.getFeatures():
                self._resultado.setdefault(feat.id(), {})["prio_bioma_amazonia"] = "Não"
            return

        if layer_biomas.crs() != self.layer.crs():
            transform = QgsCoordinateTransform(
                layer_biomas.crs(), self.layer.crs(), QgsProject.instance()
            )
            geoms_proj = []
            for g in geoms_amaz:
                try:
                    gg = QgsGeometry(g)
                    gg.transform(transform)
                    geoms_proj.append(gg)
                except Exception:
                    continue
            geoms_amaz = geoms_proj

        contador = 0
        for feat in self.layer.getFeatures():
            fid = feat.id()
            geom = feat.geometry()
            if geom is None or geom.isEmpty():
                self._resultado.setdefault(fid, {})["prio_bioma_amazonia"] = "Não"
                continue
            atende = False
            bbox = geom.boundingBox()
            for g in geoms_amaz:
                if not g.boundingBox().intersects(bbox):
                    continue
                if geom.intersects(g):
                    atende = True
                    break
            self._resultado.setdefault(fid, {})["prio_bioma_amazonia"] = "Sim" if atende else "Não"
            if atende:
                contador += 1
        self.log(f"   A8 — Bioma Amazônia: {contador} imóveis atendem")

    # ------------------------------------------------------------------ #
    # Critérios do Provedor (P1, P2, P3) — planilha de candidatos        #
    # ------------------------------------------------------------------ #

    def _carregar_planilha_candidatos(self) -> Dict[str, Dict[str, str]]:
        """Carrega a planilha CSV/XLSX e retorna um dict por CPF normalizado."""
        caminho = (getattr(self.config, "planilha_candidatos", "") or "").strip()
        if not caminho or not os.path.exists(caminho):
            return {}

        mapa_col = getattr(self.config, "mapeamento_candidatos", {}) or {}
        col_cpf = mapa_col.get("cpf")
        if not col_cpf:
            self.log("⚠ Provedor: planilha sem coluna de CPF mapeada — pulando P1/P2/P3")
            return {}

        ext = os.path.splitext(caminho)[1].lower()
        candidatos: Dict[str, Dict[str, str]] = {}

        try:
            if ext == ".csv":
                # Detectar delimitador
                with open(caminho, "r", encoding="utf-8-sig", newline="") as f:
                    sample = f.read(4096)
                    f.seek(0)
                    try:
                        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                    except Exception:
                        class _D(csv.excel):
                            delimiter = ";"
                        dialect = _D
                    reader = csv.DictReader(f, dialect=dialect)
                    for row in reader:
                        cpf = self._normalizar_cpf(row.get(col_cpf, ""))
                        if cpf:
                            candidatos[cpf] = {k: (v or "").strip() for k, v in row.items()}
            elif ext in (".xlsx", ".xls"):
                # Tentar usar openpyxl
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    self.log("⚠ Provedor: openpyxl não disponível para ler XLSX — pulando")
                    return {}
                wb = load_workbook(caminho, read_only=True, data_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                headers = [str(c) if c is not None else "" for c in next(rows_iter)]
                for row in rows_iter:
                    if not row:
                        continue
                    row_dict = {
                        headers[i]: ("" if v is None else str(v)).strip()
                        for i, v in enumerate(row[: len(headers)])
                    }
                    cpf = self._normalizar_cpf(row_dict.get(col_cpf, ""))
                    if cpf:
                        candidatos[cpf] = row_dict
            else:
                self.log(f"⚠ Provedor: extensão não suportada: {ext}")
                return {}
        except Exception as e:
            self.log(f"⚠ Provedor: erro lendo planilha — {e}")
            return {}

        self.log(f"   Planilha de candidatos: {len(candidatos)} CPFs carregados")
        return candidatos

    @staticmethod
    def _normalizar_cpf(valor) -> str:
        if valor is None:
            return ""
        s = "".join(c for c in str(valor) if c.isdigit())
        return s.zfill(11) if s else ""

    @staticmethod
    def _eh_sim(valor: str) -> bool:
        if not valor:
            return False
        v = valor.strip().lower()
        return v in {"1", "sim", "s", "yes", "y", "true", "t", "verdadeiro"}

    def _executar_provedor(self):
        """P1, P2, P3 — cruzar CPF com planilha de candidatos."""
        candidatos = self._carregar_planilha_candidatos()
        # Inicializar todos como "Não"
        for feat in self.layer.getFeatures():
            self._resultado.setdefault(feat.id(), {}).update({
                "prio_caf": "Não",
                "prio_sexo_feminino": "Não",
                "prio_sociobio": "Não",
            })

        if not candidatos:
            return

        mapa_col = getattr(self.config, "mapeamento_candidatos", {}) or {}
        col_caf = mapa_col.get("caf")
        col_sexo = mapa_col.get("sexo")
        col_sociobio = mapa_col.get("sociobio")

        # Identificar coluna de CPF do imóvel
        campo_cpf_imovel = self._identificar_campo([
            "cpf_cnpj", "cpf", "cnpj_cpf", "documento", "doc"
        ])
        if not campo_cpf_imovel:
            self.log("⚠ Provedor: imóveis sem coluna de CPF — pulando P1/P2/P3")
            return

        contadores = {"P1": 0, "P2": 0, "P3": 0}
        for feat in self.layer.getFeatures():
            cpf = self._normalizar_cpf(feat.attribute(campo_cpf_imovel))
            if not cpf or cpf not in candidatos:
                continue
            row = candidatos[cpf]
            res = self._resultado.setdefault(feat.id(), {})

            if col_caf and self._eh_sim(row.get(col_caf, "")):
                res["prio_caf"] = "Sim"
                contadores["P1"] += 1
            if col_sexo:
                sexo = (row.get(col_sexo, "") or "").strip().lower()
                if sexo in {"f", "feminino", "mulher"}:
                    res["prio_sexo_feminino"] = "Sim"
                    contadores["P2"] += 1
            if col_sociobio and self._eh_sim(row.get(col_sociobio, "")):
                res["prio_sociobio"] = "Sim"
                contadores["P3"] += 1

        self.log(
            f"   Provedor — P1: {contadores['P1']}  P2: {contadores['P2']}  P3: {contadores['P3']}"
        )

    # ------------------------------------------------------------------ #
    # Score, ranking e gravação                                           #
    # ------------------------------------------------------------------ #

    def _calcular_pct_rvn(self, feat) -> float:
        """% RVN / área imóvel (desempate)."""
        rvn = feat.attribute("RVN_area") if self._campo_existe("RVN_area") else 0
        try:
            rvn_val = float(rvn) if rvn not in (None, "") else 0.0
        except (TypeError, ValueError):
            rvn_val = 0.0
        area = self._area_imovel.get(feat.id())
        if area is None:
            area_attr = feat.attribute("area_car") if self._campo_existe("area_car") else 0
            try:
                area = float(area_attr) if area_attr not in (None, "") else 0.0
            except (TypeError, ValueError):
                area = 0.0
        if area <= 0:
            return 0.0
        return rvn_val / area

    def _aplicar_resultados(self):
        """Grava as colunas de priorização na camada (apenas para imóveis ELEGÍVEIS)."""
        # Pesos
        pesos = {
            "prio_mun_prioritario": 1, "prio_mun_controle": 1,
            "prio_mun_uniao": 1, "prio_biodiversidade": 1,
            "prio_entorno_uc": 1, "prio_apa_rppn": 1,
            "prio_entorno_ti": 1, "prio_bioma_amazonia": 1,
            "prio_caf": 3, "prio_sexo_feminino": 3, "prio_sociobio": 3,
        }

        # Mapear elegibilidade para decidir quem participa do ranking
        scores: List[Tuple[int, int, float]] = []  # (fid, score, pct_rvn) só elegíveis

        self.layer.startEditing()
        for feat in self.layer.getFeatures():
            fid = feat.id()
            res = self._resultado.get(fid, {})

            # Calcular score
            score = 0
            for col, peso in pesos.items():
                if res.get(col) == "Sim":
                    score += peso
            res["score_priorizacao"] = score

            pct_rvn = self._calcular_pct_rvn(feat)
            res["pct_rvn_total"] = round(pct_rvn, 4)

            # Verificar se imóvel é elegível
            elegivel_str = ""
            if self._campo_existe("elegibilidade"):
                elegivel_str = str(feat.attribute("elegibilidade") or "").strip().lower()
            participa = elegivel_str.startswith("eleg")

            if participa:
                scores.append((fid, score, pct_rvn))

            # Gravar todas as colunas
            for col, _, _ in CAMPOS_PRIORIZACAO:
                if col == "ranking":
                    continue  # ranking é gravado depois
                valor = res.get(col)
                if valor is None:
                    valor = "Não" if col.startswith("prio_") else 0
                idx = self.layer.fields().indexOf(col)
                if idx >= 0:
                    self.layer.changeAttributeValue(fid, idx, valor)

        # Calcular ranking entre os elegíveis
        scores.sort(key=lambda x: (-x[1], -x[2]))
        ranking_map = {fid: pos + 1 for pos, (fid, _, _) in enumerate(scores)}

        idx_ranking = self.layer.fields().indexOf("ranking")
        if idx_ranking >= 0:
            for feat in self.layer.getFeatures():
                self.layer.changeAttributeValue(
                    feat.id(), idx_ranking, ranking_map.get(feat.id(), 0)
                )

        self.layer.commitChanges()

        if scores:
            self.log(
                f"   Ranking calculado para {len(scores)} imóveis elegíveis "
                f"(score máximo: {scores[0][1]})"
            )

    # ------------------------------------------------------------------ #
    # Execução principal                                                  #
    # ------------------------------------------------------------------ #

    def executar(self) -> bool:
        """Executa toda a análise de priorização. Retorna True se concluiu."""
        try:
            self.log("=" * 60)
            self.log("INICIANDO ANÁLISE DE PRIORIZAÇÃO")
            self.log("=" * 60)

            if not self.layer or self.layer.featureCount() == 0:
                self.log("⚠ Camada vazia — priorização não executada")
                return False

            self._adicionar_campos()

            # Inicializar dicionário de resultados para todas as features
            for feat in self.layer.getFeatures():
                self._resultado.setdefault(feat.id(), {})

            # Critérios de área
            self.log("\n14.1 Critérios de municípios (A1, A2, A3, A4)...")
            self._executar_criterios_municipios()

            self.log("\n14.2 Critérios espaciais de UC e TI (A5, A6, A7)...")
            self._executar_uc_ti()

            self.log("\n14.3 Bioma Amazônia (A8)...")
            self._executar_bioma_amazonia()

            self.log("\n14.4 Critérios do provedor (P1, P2, P3)...")
            self._executar_provedor()

            self.log("\n14.5 Calculando score e ranking...")
            self._aplicar_resultados()

            self.log("=" * 60)
            self.log("ANÁLISE DE PRIORIZAÇÃO CONCLUÍDA")
            self.log("=" * 60)
            return True

        except Exception as e:
            self.log(f"❌ ERRO NA PRIORIZAÇÃO: {e}")
            traceback.print_exc()
            return False
